"""
分销缺口计算脚本
读取直营门店销量和分销标准两个Excel，计算每个门店的分销缺口，输出data.json
"""
import json
import re
import os
from collections import defaultdict
from datetime import date
from copy import copy

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# 脚本所在目录
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 门店类型映射：销量表类型 → 分销标准列名
STORE_TYPE_MAP = {
    '商超A': '商超AB',
    '商超B': '商超AB',
    '农贸菜场店': '农贸店',
    '生鲜店': '生鲜店',
    '小超市': '小超市',
    '粮油店': '粮油店',
    '食杂店': '食杂店',
    '便利店': '便利店',
    '流通批发': None,
    '批发流通渠道': None,
    '其它客户': None,
}

# 分销标准列名 → 列字母（在分销标准sheet中）
STD_COL_MAP = {
    '商超AB': 'C',
    '生鲜店': 'D',
    '小超市': 'E',
    '农贸店': 'F',
    '粮油店': 'G',
    '食杂店': 'H',
    '便利店': 'I',
}


def extract_core_name(name):
    """
    提取产品核心名（品牌+品类），去掉规格/包装/工艺等级等附属信息。
    处理顺序很重要：先清外围杂物，再去规格数字。
    """
    if name is None:
        return None
    name = str(name).strip()
    # 统一乘号
    name = name.replace('×', '*')
    name = re.sub(r'([0-9A-Za-z])x([0-9])', r'\1*\2', name)
    name = re.sub(r'([0-9])X([0-9])', r'\1*\2', name)
    # 去掉所有括号及其内容（中英文，含混合配对）
    name = re.sub(r'\([^)]*\)', '', name)
    name = re.sub(r'（[^）]*）', '', name)
    name = re.sub(r'\([^)]*）', '', name)   # 混合: ( ... ）
    name = re.sub(r'（[^）]*\)', '', name)   # 混合: （ ... )
    name = re.sub(r'[()（）]', '', name)
    # 去掉 *数字 包装数量
    name = re.sub(r'\*\d+', '', name)
    # 去掉末尾已知的描述词
    TRAILING_WORDS = [
        '圆瓶', '壶嘴装', '扁瓶装', '赠装', '加赠装', '原条码加赠装',
        '新升级', '区域限定', '黄桶', '新版',
        '独-O', 'S', 'LS', 'ZH', 'Y',
        '酱油Y', '酱油',
    ]
    for w in sorted(TRAILING_WORDS, key=len, reverse=True):
        if name.upper().endswith(w.upper()):
            name = name[:-len(w)].strip()
    # 去掉末尾 -xxx 后缀（如 -1.0, -2.25椭圆, -优质散米, -普通散米）
    name = re.sub(r'-\d+\.?\d*(?:椭圆|圆|方)?\s*$', '', name)
    name = re.sub(r'-[^/]*?(?:散米|椭圆|贸易|壶嘴|加赠)\s*$', '', name)
    # 按 / 拆分，过滤掉工艺/等级/系列等修饰词
    SKIP_TOKENS = {
        '压榨', '压', '非转', '转', '一级', '二级', '三级', '四级',
        '健康系列', '风味系列', '发酵面系列', '波纹面系列',
        '经典纸包系列', '专供纸包系列', '纸包系列',
        'MES', '金标', '非转压榨', '非转/压榨',
        '壹级', '贰级', '叁级',
        '贸易', '升级版', '定制', '活动版', '非卖品', '非卖品',
        '农贸', '纸箱', '瓶装',
    }
    parts = name.split('/')
    filtered = []
    for p in parts:
        p2 = p.strip()
        if p2 in SKIP_TOKENS:
            continue
        if re.match(r'^\d+\.?\d*(?:度|圆|方|空心)?$', p2):
            continue
        filtered.append(p)
    name = '/'.join(filtered)
    # 去掉末尾的规格数字
    name = re.sub(r'[/\s]?\d+\.?\d*\s*(?:ML|L|KG|G|克|毫升|升)\s*$', '', name, flags=re.IGNORECASE).strip()
    # 去掉末尾残留的质量/版本标记（金标、非转等）
    name = re.sub(r'[/\s]?(?:金标|非转|金标非转)\s*$', '', name).strip()
    # 去掉末尾残留的 / * - 分隔符
    name = name.rstrip('/*- ').strip()
    # 统一大小写
    name = name.upper()
    return name


def read_distribution_standard(filepath):
    """
    读取分销标准sheet，返回：
    - products: [{name, line, requirements: {store_type: value}}]
    - groups: [{store_type, rows, n, label}]  N选1分组
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb['分销标准']

    # 获取合并单元格信息
    merged = {}
    for mr in ws.merged_cells.ranges:
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                merged[(row, col)] = mr

    # 读取表头 Row2 → 列映射
    col_map = {}  # col_idx → store_type
    for c in range(3, 10):  # C=3 to I=9
        header = ws.cell(row=2, column=c).value
        if header:
            col_map[c] = str(header).strip()

    # 找到数据起始行（Row3）
    products = []
    groups = []  # N选1分组列表

    # 当前产线（从合并单元格获取）
    current_line = None

    for r in range(3, ws.max_row + 1):
        # 获取产线（列A，可能有合并）
        line_cell = ws.cell(row=r, column=1)
        line_val = line_cell.value
        if line_val and not isinstance(line_cell, MergedCell):
            current_line = str(line_val).strip()

        # 获取产品名（列B）
        product_cell = ws.cell(row=r, column=2)
        product_name = product_cell.value if not isinstance(product_cell, MergedCell) else None

        if not product_name:
            continue

        product_name = str(product_name).strip()

        # 跳过非产品条目
        skip_keywords = ['不做强制要求', '散米桶', '设置建议']
        if any(kw in product_name for kw in skip_keywords):
            continue

        # 读取各门店类型的要求值
        requirements = {}
        for col_idx, store_type in col_map.items():
            cell = ws.cell(row=r, column=col_idx)
            val = cell.value
            if val is not None:
                val = str(val).strip()
            requirements[store_type] = val

        products.append({
            'name': product_name,
            'line': current_line or '',
            'row': r,
            'requirements': requirements,
        })

    wb.close()

    # 识别N选1分组：扫描合并单元格中的 "2选1"/"3选1"/"二选一"
    for mr in ws.merged_cells.ranges:
        # 只关心数据区域（Row >= 3, Col C-I = 3-9）的合并
        if mr.min_row < 3:
            continue
        if mr.min_col < 3 or mr.min_col > 9:
            continue

        # 获取合并单元格的值
        val = ws.cell(row=mr.min_row, column=mr.min_col).value
        if val is None:
            continue
        val = str(val).strip()

        # 判断是否为N选1
        n_choose = None
        if '二选一' in val or val == '2选1':
            n_choose = 2
        elif val == '3选1':
            n_choose = 3
        else:
            continue

        store_type = col_map.get(mr.min_col)
        if not store_type:
            continue

        # 获取该组内的产品
        group_products = []
        for r in range(mr.min_row, mr.max_row + 1):
            pname = ws.cell(row=r, column=2).value
            if pname:
                group_products.append(str(pname).strip())

        if group_products:
            groups.append({
                'store_type': store_type,
                'n': n_choose,
                'rows': (mr.min_row, mr.max_row),
                'products': group_products,
            })

    wb.close()
    return products, groups


def read_store_sales(filepath):
    """
    读取门店销量表，返回：
    - stores: [{name, type, products: set, manager}]
    按门店名聚合产品
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    # 读表头行（Row 1），建立列名→列索引映射
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            headers[str(val).strip()] = c

    # 按名称匹配所需列
    def find_col(*names):
        for n in names:
            if n in headers:
                return headers[n]
        return None

    col_store = find_col('标准客户名称')
    col_type = find_col('标准客户类型(整合)')
    col_product = find_col('物料名称')
    col_manager = find_col('负责人员')

    if not all([col_store, col_type, col_product, col_manager]):
        missing = []
        if not col_store: missing.append('标准客户名称')
        if not col_type: missing.append('标准客户类型(整合)')
        if not col_product: missing.append('物料名称')
        if not col_manager: missing.append('负责人员')
        raise ValueError(f'未找到列: {", ".join(missing)}，表头为: {list(headers.keys())}')

    # 按门店聚合数据
    store_data = {}  # key: (store_name, store_type, manager) → set of product names

    for r in range(2, ws.max_row + 1):
        store_name = ws.cell(row=r, column=col_store).value
        store_type = ws.cell(row=r, column=col_type).value
        product = ws.cell(row=r, column=col_product).value
        manager = ws.cell(row=r, column=col_manager).value

        if not all([store_name, store_type, product]):
            continue

        store_name = str(store_name).strip()
        store_type = str(store_type).strip()
        product = str(product).strip()
        manager = str(manager).strip() if manager else '未知'

        key = (store_name, store_type, manager)
        if key not in store_data:
            store_data[key] = set()
        store_data[key].add(product)

    wb.close()

    # 转为列表
    stores = []
    for (name, stype, manager), products in store_data.items():
        stores.append({
            'name': name,
            'type': stype,
            'manager': manager,
            'products': products,
        })

    return stores


def find_file(*keywords):
    """在脚本目录下找包含所有关键词的 xlsx 文件"""
    for f in os.listdir(BASE_DIR):
        if f.endswith('.xlsx') and all(k in f for k in keywords):
            return os.path.join(BASE_DIR, f)
    return None

def build_data():
    sales_file = find_file('销量')
    std_file = find_file('概况')

    if not sales_file:
        raise FileNotFoundError('未找到销量表（文件名需含"销量"）')
    if not std_file:
        raise FileNotFoundError('未找到概况表（文件名需含"概况"）')

    print(f'销量表: {os.path.basename(sales_file)}')
    print(f'标准表: {os.path.basename(std_file)}')
    print('读取分销标准...')
    std_products, std_groups = read_distribution_standard(std_file)
    print(f'  分销标准产品数: {len(std_products)}')
    print(f'  N选1分组数: {len(std_groups)}')

    # 建立分销标准产品查找表（核心名 → 原始标准产品记录列表）
    # 一个核心名可能对应多个标准产品（如同一产品的不同规格）
    std_by_core = {}
    for p in std_products:
        n = extract_core_name(p['name'])
        if n:
            if n not in std_by_core:
                std_by_core[n] = []
            std_by_core[n].append(p)

    print('读取门店销量...')
    stores = read_store_sales(sales_file)
    print(f'  门店数: {len(stores)}')

    # 产品名匹配统计
    all_sales_products = set()
    for s in stores:
        all_sales_products.update(s['products'])

    print(f'  销量表产品种类: {len(all_sales_products)}')

    # 建立匹配映射：销量产品名 → 匹配到的标准产品名列表
    product_match = {}  # sales_product → [std_product_names]
    match_count = 0
    for sp in all_sales_products:
        core = extract_core_name(sp)
        if core and core in std_by_core:
            product_match[sp] = [p['name'] for p in std_by_core[core]]
            match_count += 1

    print(f'  产品名匹配: {match_count}/{len(all_sales_products)} ({100*match_count/len(all_sales_products):.1f}%)')

    # 汇总到负责人
    manager_stores = defaultdict(list)

    for store in stores:
        stype = store['type']
        mapped_type = STORE_TYPE_MAP.get(stype)

        # 构建门店当前已售产品（标准名集合）
        current_std_products = set()
        for p in store['products']:
            matched_list = product_match.get(p)
            if matched_list:
                current_std_products.update(matched_list)

        # 计算缺口
        gaps = []
        if mapped_type is not None:
            # 有分销标准：检查每个"1"标记的产品
            for sp in std_products:
                req = sp['requirements'].get(mapped_type)
                if req is None or req == 'None' or req == '':
                    continue
                if req == '1':
                    # 必须分销
                    if sp['name'] not in current_std_products:
                        gaps.append({
                            'productLine': sp['line'],
                            'requirement': '必须',
                            'product': sp['name'],
                        })

            # 检查N选1分组
            for g in std_groups:
                if g['store_type'] != mapped_type:
                    continue
                # 检查组内是否有任一产品被覆盖
                covered = any(p in current_std_products for p in g['products'])
                if not covered:
                    # 获取组内第一个产品的产线
                    first_line = ''
                    for sp in std_products:
                        if sp['name'] == g['products'][0]:
                            first_line = sp['line']
                            break
                    gaps.append({
                        'productLine': first_line,
                        'requirement': f'{g["n"]}选1',
                        'group': g['products'],
                    })

        # 按产线排序缺口
        gaps.sort(key=lambda g: (g.get('productLine', '') or '', g.get('product', '') or ''))

        manager_stores[store['manager']].append({
            'name': store['name'],
            'type': stype,
            'hasStandard': mapped_type is not None,
            'gaps': gaps,
        })

    # 构建输出
    managers = []
    for name, mstores in sorted(manager_stores.items()):
        gap_store_count = sum(1 for s in mstores if len(s['gaps']) > 0)
        managers.append({
            'name': name,
            'storeCount': len(mstores),
            'gapStoreCount': gap_store_count,
            'stores': sorted(mstores, key=lambda s: (-len(s['gaps']), s['name'])),
        })

    # 按缺口门店数降序排列
    managers.sort(key=lambda m: (-m['gapStoreCount'], m['name']))

    data = {
        'generated': date.today().isoformat(),
        'sourceSales': os.path.basename(sales_file),
        'sourceStd': os.path.basename(std_file),
        'managers': managers,
    }

    output_file = os.path.join(BASE_DIR, 'data.json')
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f'\n输出: {output_file}')
    print(f'负责人数: {len(managers)}')
    print(f'有缺口的门店数: {sum(m["gapStoreCount"] for m in managers)}')

    # 自动提交到 GitHub
    auto_git_push()


def auto_git_push():
    """自动 git add + commit + push"""
    import subprocess, sys
    repo_dir = BASE_DIR

    # 检查是否有变更
    result = subprocess.run(
        ['git', '-C', repo_dir, 'status', '--porcelain', 'data.json'],
        capture_output=True, text=True
    )
    if not result.stdout.strip():
        print('data.json 无变更，跳过提交')
        return

    print('\n提交到 GitHub...')
    cmds = [
        ['git', '-C', repo_dir, 'add', 'data.json'],
        ['git', '-C', repo_dir, 'commit', '-m', f'更新数据 {date.today().isoformat()}'],
        ['git', '-C', repo_dir, 'push'],
    ]
    for cmd in cmds:
        r = subprocess.run(cmd, capture_output=True, text=True)
        if r.returncode != 0:
            print(f'  失败: {" ".join(cmd)}')
            print(f'  {r.stderr.strip()}')
            return
    print('  已推送到 GitHub')


if __name__ == '__main__':
    build_data()
    try:
        input("按 Enter 键退出...")
    except EOFError:
        pass
